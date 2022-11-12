
import logging
import os
from os.path import exists

import db
import openpyxl
import weights as weight_server
from flask import (Flask, jsonify, make_response, render_template, request,
                   send_file)
from library import get_date_range, make_error, make_success
from werkzeug.utils import secure_filename

app = Flask(__name__)

logging.basicConfig(filename='/tmp/cilogs/b_record.log', level=logging.DEBUG, format=f'%(asctime)s %(levelname)s %(name)s %(threadName)s : %(message)s')

database = db.connect()



@app.route("/")
def home():
  return "Welcome to GanShmuel Billing Server"



@app.route("/ui")
def ui():
  return render_template("ui.html")



@app.route("/provider", methods=["POST"])
def provider_create():
  new_name = request.json.get('name')
  provider_id = db.create_provider(database, new_name)
  new_provider = db.get_provider(database, provider_id)
  return make_success({"id": new_provider[0], "name": new_provider[1]})



@app.route("/provider/<id>", methods=["PUT"])
def provider_update(id):
  new_name = request.json.get('name')
  db.update_provider(database, id, new_name)
  new_provider = db.get_provider(database, id)
  return make_success({"id": new_provider[0], "name": new_provider[1]})



@app.route("/rates", methods=["GET"])
def rates_get():
  store_path = "/in/store.txt"
  if not exists(store_path):
    return make_error(500, 'File has not yet been uploaded')

  file = None
  with open(store_path, 'r') as store_reader:
    file = store_reader.read()
  
  if file == None:
    return make_error(500, 'File has not yet been uploaded')

  file_name = os.path.join("/in", file)

  if not exists(file_name):
    return make_error(500, 'File specified does not exist')

  return send_file(file_name, as_attachment=True, mimetype='application/vnd.ms-excel', download_name=file)



@app.route("/rates", methods=["POST"])
def rates_create():
  try:
    file_name = request.json.get('file')
    file = f"/in/{file_name}"

    if not exists(file):
      return make_error(500, 'File specified does not exist')

    # store file name
    with open("/in/store.txt", 'w') as store_reader:
      store_reader.write(file_name)

    dataframe = openpyxl.load_workbook(file)
    dataframe1 = dataframe.active
    
    testArray = []
    # Iterate the loop to read the cell values
    for row in range(1, dataframe1.max_row):
      
      testDict={}
      for col in dataframe1.iter_cols(1, dataframe1.max_column):
        
        if col[0].value == "Scope":
          testDict[col[0].value] = str(col[row].value)
        elif col[0].value == "Product":
          testDict[col[0].value] = str(col[row].value)    
        else:
          testDict[col[0].value] = int( col[row].value )

      testArray.append(testDict)  
    
    for x in testArray:

      if db.get_one_rate(database, x["Product"], x["Scope"]) != []:
        db.update_rates_same_pid_scope(database, x["Product"], x["Rate"], x["Scope"])
      else:
        db.create_rates(database, x["Product"], x["Rate"], x["Scope"])

    return make_success({"status_code": 200, "success": "true"})

  except:
    return make_error(400, 'Bad request')
  
    

@app.route("/truck", methods=["POST"])
def truck_create():
  number_plate = request.json.get('number_plate')
  provider_id = request.json.get('provider_id')

  provider = db.get_provider(database, provider_id)
  if provider == None:
    return make_error(404, "Provider ID does not exist")

  truck = db.get_truck(database, number_plate)
  if truck != None:
    return make_error(404, "Truck ID already exist")

  db.create_truck(database, number_plate, provider_id)
  truck = db.get_truck(database, number_plate)
  return make_success({"id": truck[0], "provider_id": truck[1]})



@app.route("/truck/<id>", methods=["PUT"])
def truck_update(id):
  provider_id = request.json.get('provider_id')

  truck = db.get_truck(database, id)
  if truck == None:
    return make_error(404, "Truck ID does not exist")

  provider = db.get_provider(database, provider_id)
  if provider == None:
    return make_error(404, "Provider ID does not exist")

  db.update_truck(database, id, provider_id)
  truck = db.get_truck(database, id)
  return make_success({"id": truck[0], "provider_id": truck[1]})



@app.route("/truck/<id>", methods=["GET"])
def truck_get(id):
  _from = request.args.get('from')
  _to = request.args.get('to')
  
  dates = get_date_range(_from, _to)

  # check db first
  truck = db.get_truck(database, id)
  if not truck:
    return make_error(404, "Truck ID does not exist")

  try:
    truck_data = weight_server.get_item(id, dates)
    return make_success(truck_data)
  except Exception as ex:
    return make_error(500, str(ex))



@app.route("/bill/<provider_id>", methods=["GET"])
def bill_get(provider_id):
  try:
    _from = request.args.get('from')
    _to = request.args.get('to')
    dates = get_date_range(_from, _to)

    # print(f"dates -> {dates}")
    
    # find provider
    provider = db.get_provider(database, provider_id)
    # print(f"provider -> {provider}")
    if provider == None:
      return make_error(404, 'Provider does not exist')

    # get trucks
    trucks = db.get_truck_for_provider(database, provider_id)
    # print(f"trucks -> {trucks}")

    #  get session IDs for trucks
    sessions = []
    for truck in trucks:
      try:
        truck_data = weight_server.get_item(truck[0], dates)
        for sess in truck_data['sessions']:
          sessions.append(sess)
      except:
        pass
      
    # get weights, and filter sessions
    weights = weight_server.get_weight(dates)

    products = {}
    # print(f"weight -> {weights}")

    totals = 0
    for weight in weights:

      if weight['id'] in sessions:
        # get session data for weight
        try:
          session = weight_server.get_session(weight['id'])
          # print(f"session -> {session}")
          
          rate = db.get_rate_for_product(database, provider_id, weight['produce'])
          # print(f"rate -> {rate}")

          if weight['produce'] in products.keys():
            products[weight['produce']]['count'] += 1
          else:
            prod = {
              "product": weight['produce'],
              "count": 1,
              "amount": session['neto'],
              "rate": rate,
              "pay": session['neto'] * rate,
            }
            products[weight['produce']] = prod
            totals += session['neto'] * rate
        except Exception as ex:
          return make_error(500, str(ex))
      

    result = {
      "id": provider_id,
      "name": provider[1],
      "from": dates['from'],
      "to": dates['to'],
      "truckCount": len(trucks),
      "sessionCount": len(sessions),
      "products": list(products.values()),
      "total": totals
    }

    return make_success(result)
    
  except Exception as e:
    return make_error(500, str(e))

@app.route("/health", methods=["GET"])
def health():
  if db.connect() != None:
    return make_response("OK", 200)
    
  return make_response("Failure", 500)

if __name__ == '__main__':
  app.run(host='0.0.0.0', port=5000, debug=True) # ! TODO remove debug
  app.logger.info('Server is Working')
