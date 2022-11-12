
import logging
import os

import db
import openpyxl
import weights as weight_server
from flask import (Flask, jsonify, make_response, render_template, request,
                   send_file)
from library import get_date_range, make_error, make_success
from werkzeug.utils import secure_filename

app = Flask(__name__)
logging.basicConfig(filename='/docker-entrypoint-initdb.d/tmp/record.log', level=logging.DEBUG, format=f'%(asctime)s %(levelname)s %(name)s %(threadName)s : %(message)s')

database = db.connect()


@app.route("/")
def home():
  return render_template("index.html")

@app.route("/test_route")
def test_route():
  return jsonify({'message': 'Hello, world!'})

@app.route("/provider", methods=["POST"])
def provider_create():
  new_name = request.json.get('name')
  provider_id = db.create_provider(database, new_name)
  return make_success({"id": provider_id})

@app.route("/provider/<id>", methods=["PUT"])
def provider_update(id):
  new_name = request.json.get('name')
  db.update_provider(database, id, new_name)
  return make_success({"id": id, "name": new_name})



@app.route("/rates", methods=["GET"])
def rates_get():
  dir_name = os.path.join("/in", 'rates.xlsx')

  return send_file(dir_name, mimetype='xlsx')



@app.route("/rates", methods=["POST"])
def rates_create():
  try:
    dataframe = openpyxl.load_workbook("/in/rates.xlsx")
    dataframe1 = dataframe.active
    
    testArray = []
    # Iterate the loop to read the cell values
    for row in range(1, dataframe1.max_row):
      
      testDict={}
      for col in dataframe1.iter_cols(1, dataframe1.max_column):
        
        if col[0].value == "Scope":
          testDict[col[0].value] = str(col[row].value)
        elif col[0].value == "Product":
          testDict[col[0].value] = str(col[row].value)    
        else:
          testDict[col[0].value] = int( col[row].value )

      testArray.append(testDict)  
    
    for x in testArray:

      if db.get_one_rate(database, x["Product"], x["Scope"]) != []:
        db.update_rates_same_pid_scope(database, x["Product"], x["Rate"], x["Scope"])
      else:
        db.create_rates(database, x["Product"], x["Rate"], x["Scope"])

    return make_success({"status_code": 200, "success": "true"})

  except:
    return make_error(400, 'Bad request')
  
    


@app.route("/truck", methods=["POST"])
def truck_create():
  number_plate = request.json.get('number_plate')
  provider_id = request.json.get('provider_id')
  truck = db.create_truck(database, number_plate, provider_id)
  return make_success({"id": truck})

@app.route("/truck/<id>", methods=["PUT"])
def truck_update(id):
  provider_id = request.json.get('provider_id')
  
  truck = db.update_truck(database, id, provider_id)
  
  return make_success({"id": truck, "provider_id": provider_id})

@app.route("/truck/<id>", methods=["GET"])
def truck_get(id):
  _from = request.args.get('from')
  _to = request.args.get('to')
  
  dates = get_date_range(_from, _to)

  # check db first
  truck = db.get_truck(database, id)

  if not truck:
    return make_error(404, 'not found')

  truck_data = weight_server.get_item(id, dates)

  if not truck_data:
    return make_error(404, 'not found')

  return make_success(truck_data)

@app.route("/bill/<provider_id>", methods=["GET"])
def bill_get(provider_id):
  try:
    _from = request.args.get('from')
    _to = request.args.get('to')
    dates = get_date_range(_from, _to)

    print(f"dates -> {dates}")
    
    # find provider
    provider = db.get_provider(database, provider_id)
    print(f"provider -> {provider}")
    if provider == None:
      return make_error(400, 'provider does not exist')

  # # ! test
    # try:
    #   return weight_server.get_item(provider_id, dates)
    # except:
    #   pass
  #   rate = db.get_rate_for_product(database, '10044', 'Mandarin')
  #   print(f"rate -> {rate}")
  #   return make_success(rate)
  # # ! test

    # get trucks
    trucks = db.get_truck_for_provider(database, provider_id)
    print(f"trucks -> {trucks}")

    #  get session IDs for trucks
    sessions = []
    for truck in trucks:
      truck_data = weight_server.get_item(truck[0], dates)
      for sess in truck_data['sessions']:
        sessions.append(sess)


    # get weights, and filter sessions
    weights = weight_server.get_weight(dates)

    products = {}
    print(f"sess -> {weights}")
    for sess in weights:

      if sess['id'] in sessions:
        rate = db.get_rate_for_product(database, provider_id, sess['produce'])

        # if products[sess['produce']] == None:
        if sess['produce'] in products.keys():
          products[sess['produce']]['count'] += 1
        else:
          prod = {
            "product": sess['produce'],
            "count": 1,
            "amount": sess['neto'],
            "rate": rate,
            "pay": sess['neto'] * rate,
          }
          products[sess['produce']] = prod
      

    result = {
      "id": provider_id,
      "name": provider[1],
      "from": dates['from'],
      "to": dates['to'],
      "truckCount": len(trucks),
      "sessionCount": len(sessions),
      "products": list(products.values()),
      "total": 0 # ! TODO
    }

    return make_success(result)
    
  except Exception as e:
    return make_error(500, str(e))

@app.route("/health", methods=["GET"])
def health():
  if db.connect() != None:
    return make_response("OK", 200)
    
  return make_response("Failure", 500)

if __name__ == '__main__':
  app.run(host='0.0.0.0', port=5000)
  app.logger.info('Server is Working')
